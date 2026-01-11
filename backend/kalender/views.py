from rest_framework import viewsets, permissions, serializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.http import HttpResponse
from .models import Gemeindetermin, Mitarbeitereintrag, Raumbelegung, Raum, KalenderKategorie, MitarbeiterKategorie, Mitarbeiter
from .serializers import GemeindeterminSerializer, MitarbeitereintraginSerializer, RaumbelegungSerializer, RaumSerializer, KalenderKategorieSerializer, MitarbeiterKategorieSerializer, MitarbeiterSerializer
from .services import FeiertagService
from datetime import datetime, timedelta
import calendar


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def feiertage_view(request):
    """API-Endpoint für Feiertage"""
    jahr = request.GET.get('jahr')
    monat = request.GET.get('monat')
    
    if not jahr:
        from datetime import date
        jahr = date.today().year
    else:
        jahr = int(jahr)
    
    if monat:
        monat = int(monat)
        feiertage = FeiertagService.get_feiertage_for_month(jahr, monat)
    else:
        feiertage = FeiertagService.get_feiertage(jahr)
    
    return Response({
        'jahr': jahr,
        'monat': monat,
        'feiertage': feiertage
    })


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def gemeindetermine_ics_export(request):
    """Export Gemeindetermine als ICS Datei"""
    jahr = request.GET.get('jahr')
    monat = request.GET.get('monat')
    
    termine = Gemeindetermin.objects.all()
    
    if jahr:
        termine = termine.filter(datum__year=int(jahr))
    if monat:
        termine = termine.filter(datum__month=int(monat))
    
    termine = termine.order_by('datum', 'startzeit')
    
    # ICS Datei erstellen
    ics_content = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//FECG Lahr//Gemeindekalender//DE", "CALSCALE:GREGORIAN", "METHOD:PUBLISH"]
    
    for termin in termine:
        # Kombiniere Datum und Zeit für DTSTART/DTEND
        dtstart = datetime.combine(termin.datum, termin.startzeit)
        dtend = datetime.combine(termin.datum, termin.endzeit)
        
        # UID generieren
        uid = f"gemeindetermin-{termin.id}@fecg-lahr.de"
        
        # Kategorie-Info
        kategorie_name = ""
        if termin.kategorie_neu:
            kategorie_name = termin.kategorie_neu.bezeichnung
        elif termin.kategorie:
            kategorie_name = termin.get_kategorie_display()
        
        ics_content.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}",
            f"DTSTART:{dtstart.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{dtend.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:{termin.titel}",
            f"DESCRIPTION:{termin.beschreibung or ''}",
            f"CATEGORIES:{kategorie_name}",
            "END:VEVENT"
        ])
    
    ics_content.append("END:VCALENDAR")
    
    response = HttpResponse("\r\n".join(ics_content), content_type='text/calendar; charset=utf-8')
    filename = f"gemeindetermine_{jahr or 'alle'}_{monat or 'alle'}.ics"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def jahreskalender_excel_export(request):
    """Export Jahreskalender als Excel"""
    jahr = request.GET.get('jahr')
    if not jahr:
        jahr = datetime.now().year
    else:
        jahr = int(jahr)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return Response({"error": "openpyxl nicht installiert"}, status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Kalender {jahr}"
    
    # Header
    ws['A1'] = f"Jahreskalender {jahr}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')
    
    # Monats-Übersicht
    row = 3
    for monat in range(1, 13):
        monat_name = calendar.month_name[monat]
        ws[f'A{row}'] = monat_name
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        
        row += 1
        
        # Header für Termine
        ws[f'A{row}'] = "Datum"
        ws[f'B{row}'] = "Titel"
        ws[f'C{row}'] = "Zeit"
        ws[f'D{row}'] = "Kategorie"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        row += 1
        
        # Termine für diesen Monat
        termine = Gemeindetermin.objects.filter(
            datum__year=jahr,
            datum__month=monat
        ).order_by('datum', 'startzeit')
        
        feiertage = FeiertagService.get_feiertage_for_month(jahr, monat)
        
        for termin in termine:
            ws[f'A{row}'] = termin.datum.strftime('%d.%m.%Y')
            ws[f'B{row}'] = termin.titel
            ws[f'C{row}'] = f"{termin.startzeit.strftime('%H:%M')} - {termin.endzeit.strftime('%H:%M')}"
            if termin.kategorie_neu:
                ws[f'D{row}'] = termin.kategorie_neu.bezeichnung
            elif termin.kategorie:
                ws[f'D{row}'] = termin.get_kategorie_display()
            row += 1
        
        # Feiertage hinzufügen
        for feiertag in feiertage:
            datum_obj = datetime.strptime(feiertag['datum'], '%Y-%m-%d').date()
            ws[f'A{row}'] = datum_obj.strftime('%d.%m.%Y')
            ws[f'B{row}'] = feiertag['name']
            ws[f'C{row}'] = "-"
            ws[f'D{row}'] = "Feiertag"
            ws[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 1
        
        row += 2  # Abstand zwischen Monaten
    
    # Spaltenbreiten anpassen
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # Excel-Datei in HttpResponse schreiben
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="jahreskalender_{jahr}.xlsx"'
    
    return response


class MitarbeiterViewSet(viewsets.ModelViewSet):
    """ViewSet für Mitarbeiter - nur für Django-Admins"""
    queryset = Mitarbeiter.objects.filter(aktiv=True)
    serializer_class = MitarbeiterSerializer
    permission_classes = [permissions.IsAdminUser]


class KalenderKategorieViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet für Kalender-Kategorien (nur lesen)"""
    queryset = KalenderKategorie.objects.filter(aktiv=True)
    serializer_class = KalenderKategorieSerializer
    permission_classes = [permissions.AllowAny]


class MitarbeiterKategorieViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet für Mitarbeiter-Kategorien - nur für Django-Admins"""
    queryset = MitarbeiterKategorie.objects.filter(aktiv=True)
    serializer_class = MitarbeiterKategorieSerializer
    permission_classes = [permissions.IsAdminUser]


class GemeindeterminViewSet(viewsets.ModelViewSet):
    """ViewSet für Gemeindetermine"""
    queryset = Gemeindetermin.objects.all()
    serializer_class = GemeindeterminSerializer
    
    def get_permissions(self):
        """Lesen öffentlich, Schreiben nur authentifiziert"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(erstellt_von=self.request.user)


class MitarbeitereintraginViewSet(viewsets.ModelViewSet):
    """ViewSet für Mitarbeitereinträge - nur für Django-Admins"""
    queryset = Mitarbeitereintrag.objects.all()
    serializer_class = MitarbeitereintraginSerializer
    permission_classes = [permissions.IsAdminUser]

    def perform_create(self, serializer):
        serializer.save(erstellt_von=self.request.user)


class RaumbelegungViewSet(viewsets.ModelViewSet):
    """ViewSet für Raumbelegungen mit Überschneidungsprüfung"""
    queryset = Raumbelegung.objects.all()
    serializer_class = RaumbelegungSerializer
    
    def get_permissions(self):
        """Lesen öffentlich, Schreiben nur authentifiziert"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def perform_create(self, serializer):
        """Erstelle neue Raumbelegung mit Überschneidungsprüfung"""
        # Setze erstellt_von nur wenn User authentifiziert ist
        if self.request.user.is_authenticated:
            raumbelegung = serializer.save(erstellt_von=self.request.user)
        else:
            raumbelegung = serializer.save()
        
        # Prüfe auf Überschneidungen nach dem Speichern
        # (da wir die ManyToMany Räume brauchen)
        raum_ids = self.request.data.get('raum', [])
        if raum_ids:
            ergebnis = raumbelegung.ueberschneidung_pruefung(raum_ids)
            if not ergebnis['ok']:
                # Lösche die Buchung wieder und gebe Fehler zurück
                raumbelegung.delete()
                raise serializers.ValidationError({
                    'ueberschneidung': 'Es gibt zeitliche Überschneidungen',
                    'konflikte': ergebnis['konflikte']
                })
    
    def perform_update(self, serializer):
        """Update Raumbelegung mit Überschneidungsprüfung"""
        raumbelegung = serializer.save()
        
        # Prüfe auf Überschneidungen
        raum_ids = self.request.data.get('raum', [])
        if raum_ids:
            ergebnis = raumbelegung.ueberschneidung_pruefung(raum_ids)
            if not ergebnis['ok']:
                raise serializers.ValidationError({
                    'ueberschneidung': 'Es gibt zeitliche Überschneidungen',
                    'konflikte': ergebnis['konflikte']
                })


class RaumViewSet(viewsets.ModelViewSet):
    """ViewSet für Räume"""
    queryset = Raum.objects.all()
    serializer_class = RaumSerializer
    
    def get_permissions(self):
        """Lesen öffentlich, Schreiben nur authentifiziert"""
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]
